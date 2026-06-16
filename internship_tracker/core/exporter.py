"""
Competitor Tracker - Excel & CSV Exporter
4-sheet Excel with price-change color coding + CSV (UTF-8 BOM).
"""

import os
import csv
from datetime import datetime, timedelta
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from internship_tracker.core.config import DATA_DIR


def export_to_excel(products, new_hashes=None, price_changes=None,
                    output_path=None):
    """
    Export products to formatted Excel with 4 sheets.

    Sheet 1 "全部商品"   — all active products with color coding
    Sheet 2 "新品上架"   — products first seen in this run (new_hashes)
    Sheet 3 "价格变动明细" — products whose price changed this run
    Sheet 4 "统计概览"   — per-platform counts, price trends, stock summary

    Color coding rules (Sheet 1 only, priority order):
      1. New product (in new_hashes or first_seen < 24h ago)
         → entire row light yellow (#FFFFCC)
      2. Price increased → price cell light red (#FFCCCC)
      3. Price decreased → price cell light green (#CCFFCC)

    Args:
        products: List[dict] from ProductDatabase.export_all()
        new_hashes: set of product_hash values new in this crawl
        price_changes: dict {product_hash: (old_price, new_price)} or None
        output_path: optional custom path; auto-generated if None

    Returns:
        str: absolute path to the saved .xlsx file
    """
    new_hashes = new_hashes or set()
    price_changes = price_changes or {}
    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    cutoff_24h = (now - timedelta(hours=24)).strftime("%Y-%m-%d %H:%M:%S")

    if output_path is None:
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            DATA_DIR, f"competitor-products_{timestamp}.xlsx")

    wb = Workbook()

    # ==================================================================
    # Shared Styles
    # ==================================================================
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    # Color fills
    new_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                           fill_type="solid")       # yellow — new product
    up_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                          fill_type="solid")         # red — price up
    down_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC",
                            fill_type="solid")       # green — price down
    new_font = Font(name="微软雅黑", bold=True, color="D9001B")  # red bold status
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")

    # Column definitions
    headers = [
        "状态", "平台", "商品标题", "SKU ID", "类目", "店铺",
        "当前价格", "原价", "货币", "销量", "评价数", "评分",
        "库存状态", "上架日期", "首次监控", "最后更新", "商品链接",
    ]
    col_widths = [10, 10, 40, 16, 14, 18, 12, 12, 8, 10, 10, 8,
                  12, 14, 18, 18, 20]

    # ==================================================================
    # Helper: determine product status and color
    # ==================================================================
    def _classify(p):
        """Return (status_label, fill_or_None) for a product row."""
        h = p.get("product_hash", "")
        first_seen = p.get("first_seen", "")

        # Priority 1: new product
        is_new = (h in new_hashes) or (first_seen >= cutoff_24h)
        if is_new:
            return "新品上架", new_fill

        # Priority 2: price change
        if h in price_changes:
            old_p, new_p = price_changes[h]
            if new_p > old_p:
                return "价格上涨 ↑", up_fill
            else:
                return "价格下跌 ↓", down_fill

        return "已记录", None

    # ==================================================================
    # Sheet 1: All Products
    # ==================================================================
    ws = wb.active
    ws.title = "全部商品"
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = \
            header_fill, header_font, header_align, thin_border

    for ri, p in enumerate(products, 2):
        status, row_fill = _classify(p)
        row = [
            status,
            p.get("platform", ""),
            p.get("product_title", ""),
            p.get("sku_id", ""),
            p.get("category", ""),
            p.get("shop_name", ""),
            p.get("current_price"),
            p.get("original_price"),
            p.get("currency", "USD"),
            p.get("sales_volume", 0),
            p.get("review_count", 0),
            p.get("rating"),
            p.get("stock_status", ""),
            p.get("listing_date", ""),
            p.get("first_seen", ""),
            p.get("last_updated", ""),
            p.get("product_url", ""),
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.alignment, c.border = data_font, data_align, thin_border
            # Apply row-level fill (new product → entire row yellow)
            if row_fill:
                c.fill = row_fill
            if ci == 1 and status != "已记录":
                c.font = new_font

        # Hyperlink on product_url column (col 17)
        uc = ws.cell(row=ri, column=17)
        if p.get("product_url"):
            uc.hyperlink = p["product_url"]
            uc.font = link_font

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_col_letter(len(headers))}{len(products) + 1}"

    # ==================================================================
    # Sheet 2: New Products
    # ==================================================================
    new_products = [p for p in products if
                    p.get("product_hash", "") in new_hashes
                    or p.get("first_seen", "") >= cutoff_24h]
    if new_products:
        ws2 = wb.create_sheet(title="新品上架")
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill, c.font, c.alignment, c.border = \
                header_fill, header_font, header_align, thin_border
        for ri, p in enumerate(new_products, 2):
            row = [
                "新品上架", p.get("platform", ""),
                p.get("product_title", ""), p.get("sku_id", ""),
                p.get("category", ""), p.get("shop_name", ""),
                p.get("current_price"), p.get("original_price"),
                p.get("currency", "USD"), p.get("sales_volume", 0),
                p.get("review_count", 0), p.get("rating"),
                p.get("stock_status", ""), p.get("listing_date", ""),
                p.get("first_seen", ""), p.get("last_updated", ""),
                p.get("product_url", ""),
            ]
            for ci, val in enumerate(row, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.font, c.alignment, c.border, c.fill = \
                    data_font, data_align, thin_border, new_fill
            uc = ws2.cell(row=ri, column=17)
            if p.get("product_url"):
                uc.hyperlink = p["product_url"]
                uc.font = link_font
        for i, w in enumerate(col_widths, 1):
            ws2.column_dimensions[_col_letter(i)].width = w
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

    # ==================================================================
    # Sheet 3: Price Change Details
    # ==================================================================
    pc_products = [p for p in products
                   if p.get("product_hash", "") in price_changes]
    pc_headers = [
        "状态", "平台", "商品标题", "SKU ID", "类目", "店铺",
        "旧价格", "新价格", "涨跌额", "涨跌幅(%)", "货币",
        "销量", "评价数", "库存状态", "首次监控", "最后更新", "商品链接",
    ]
    pc_widths = [10, 10, 40, 16, 14, 18, 12, 12, 10, 12, 8,
                 10, 10, 12, 18, 18, 20]

    if pc_products:
        ws3 = wb.create_sheet(title="价格变动明细")
        for ci, h in enumerate(pc_headers, 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill, c.font, c.alignment, c.border = \
                header_fill, header_font, header_align, thin_border

        for ri, p in enumerate(pc_products, 2):
            h = p.get("product_hash", "")
            old_p, new_p = price_changes.get(h, (None, None))
            direction = "价格上涨 ↑" if (new_p or 0) > (old_p or 0) \
                else "价格下跌 ↓"
            change = (new_p or 0) - (old_p or 0)
            pct = (change / old_p * 100) if old_p else 0

            row = [
                direction, p.get("platform", ""),
                p.get("product_title", ""), p.get("sku_id", ""),
                p.get("category", ""), p.get("shop_name", ""),
                old_p, new_p, round(change, 2), round(pct, 2),
                p.get("currency", "USD"), p.get("sales_volume", 0),
                p.get("review_count", 0), p.get("stock_status", ""),
                p.get("first_seen", ""), p.get("last_updated", ""),
                p.get("product_url", ""),
            ]
            row_fill = up_fill if direction.startswith("价格上涨") \
                else down_fill
            for ci, val in enumerate(row, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.font, c.alignment, c.border = data_font, data_align, thin_border
                c.fill = row_fill
            uc = ws3.cell(row=ri, column=17)
            if p.get("product_url"):
                uc.hyperlink = p["product_url"]
                uc.font = link_font

        for i, w in enumerate(pc_widths, 1):
            ws3.column_dimensions[_col_letter(i)].width = w
        ws3.row_dimensions[1].height = 28
        ws3.freeze_panes = "A2"

    # ==================================================================
    # Sheet 4: Stats Overview
    # ==================================================================
    ws4 = wb.create_sheet(title="统计概览")
    for ci, h in enumerate(["指标", "数值"], 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = header_fill, header_font, header_align

    # Compute stats
    total = len(products)
    new_cnt = sum(1 for p in products if
                  p.get("product_hash", "") in new_hashes
                  or p.get("first_seen", "") >= cutoff_24h)
    pc_cnt = len(pc_products)
    up_cnt = sum(1 for p in pc_products
                 if price_changes.get(p.get("product_hash", ""), (0, 0))[1]
                 > price_changes.get(p.get("product_hash", ""), (0, 0))[0])
    down_cnt = pc_cnt - up_cnt
    oos_cnt = sum(1 for p in products
                  if p.get("stock_status") == "out_of_stock")

    stats_data = [
        ("导出时间", now_str),
        ("全部商品数", total),
        ("新品上架数", new_cnt),
        ("已记录商品数", total - new_cnt),
        ("价格变动商品数", pc_cnt),
        ("  └ 涨价商品数", up_cnt),
        ("  └ 降价商品数", down_cnt),
        ("缺货商品数", oos_cnt),
    ]

    # Per-platform breakdown
    platform_counter = Counter(p.get("platform", "") for p in products)
    stats_data.append(("", ""))
    stats_data.append(("--- 各平台商品分布 ---", ""))
    for plat, cnt in platform_counter.most_common():
        stats_data.append((plat or "未知平台", cnt))

    # Stock status breakdown
    stock_counter = Counter(p.get("stock_status", "") for p in products)
    stats_data.append(("", ""))
    stats_data.append(("--- 库存状态分布 ---", ""))
    for status, cnt in stock_counter.most_common():
        label = {"in_stock": "有库存", "low_stock": "低库存",
                 "out_of_stock": "缺货", "pre_order": "预售"}.get(status, status)
        stats_data.append((label, cnt))

    for ri, (k, v) in enumerate(stats_data, 2):
        ws4.cell(row=ri, column=1, value=k).font = data_font
        ws4.cell(row=ri, column=2, value=v).font = data_font
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 15

    wb.save(output_path)
    return output_path


def export_to_csv(products, output_path=None):
    """
    Export products to CSV with UTF-8 BOM for Chinese Excel compatibility.

    Preserves original CSV logic — no colors, pure data.
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            DATA_DIR, f"competitor-products_{ts}.csv")

    headers = [
        "平台", "商品标题", "SKU ID", "类目", "店铺",
        "当前价格", "原价", "货币", "销量", "评价数", "评分",
        "库存状态", "上架日期", "首次监控", "最后更新", "商品链接",
    ]
    fields = [
        "platform", "product_title", "sku_id", "category", "shop_name",
        "current_price", "original_price", "currency", "sales_volume",
        "review_count", "rating", "stock_status", "listing_date",
        "first_seen", "last_updated", "product_url",
    ]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for p in products:
            w.writerow([p.get(f, "") for f in fields])
    return output_path


def _col_letter(n):
    """Convert 1-based column index to Excel letter(s)."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
